import json, os, requests
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XlImage
from PIL import Image as PilImage

JSON_PATH = "movies.json"
XLSX_PATH = "movies_with_posters.xlsx"
IMG_DIR = "posters"
os.makedirs(IMG_DIR, exist_ok=True)

with open(JSON_PATH, encoding="utf-8") as f:
    movies = json.load(f)

session = requests.Session()
session.headers.update({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"})

COL_W = 80  # width to resize poster thumbnails

def download_poster(url, mid):
    fname = os.path.join(IMG_DIR, f"{mid}.jpg")
    if os.path.exists(fname):
        return fname
    try:
        r = session.get(url.replace("@464w_644h_1e_1c", "@160w_222h_1e_1c"), timeout=15, verify=False)
        if r.status_code != 200:
            r = session.get(url, timeout=15, verify=False)
        img = PilImage.open(BytesIO(r.content)).convert("RGB")
        w, h = img.size
        ratio = COL_W / w
        img = img.resize((COL_W, int(h * ratio)), PilImage.LANCZOS)
        img.save(fname, "JPEG", quality=80)
        print(f"  [OK] {mid}")
        return fname
    except Exception as e:
        print(f"  [FAIL] {mid}: {e}")
        return None

print(f"Downloading {len(movies)} posters...")
for m in movies:
    m["_poster"] = download_poster(m["cover"], m["id"])

# --- Excel ---
wb = Workbook()
ws = wb.active
ws.title = "電影資訊"

headers = ["編號", "海報", "片名", "中文片名", "英文片名", "評分", "分類", "地區", "片長(分鐘)", "上映日期", "導演", "演員"]
hdr_font = Font(name="Microsoft JhengHei", bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor="1A1A2E")
hdr_align = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin", color="D0D5DD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for ci, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=ci, value=h)
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align; c.border = border

body_font = Font(name="Microsoft JhengHei", size=10)
wrap = Alignment(vertical="center", wrap_text=True)
center = Alignment(horizontal="center", vertical="center")

ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 34
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 28
ws.column_dimensions["F"].width = 6
ws.column_dimensions["G"].width = 16
ws.column_dimensions["H"].width = 14
ws.column_dimensions["I"].width = 10
ws.column_dimensions["J"].width = 14
ws.column_dimensions["K"].width = 14
ws.column_dimensions["L"].width = 36

ws.row_dimensions[1].height = 22

for ri, m in enumerate(movies, 2):
    parts = m["name"].split(" - ", 1)
    cn = parts[0]
    en = parts[1] if len(parts) > 1 else ""
    cats = "、".join(m.get("categories", []))
    dirs = "、".join(m.get("directors", []))
    actors = "、".join(a["name"] for a in m.get("actors", []))
    duration = m.get("duration", "").replace(" 分钟", "")
    release = m.get("release", "").replace(" 上映", "")

    row_data = [m["id"], "", m["name"], cn, en, m["score"], cats, m.get("region", ""), duration, release, dirs, actors]
    for ci, val in enumerate(row_data, 1):
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.font = body_font
        cell.alignment = center if ci in (1, 6, 9) else wrap
        cell.border = border

    # insert poster
    if m["_poster"]:
        img = XlImage(m["_poster"])
        img.width = 70
        img.height = 98
        ws.add_image(img, f"B{ri}")
        ws.row_dimensions[ri].height = 100

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:L{len(movies)+1}"
wb.save(XLSX_PATH)
print(f"\nDone -> {XLSX_PATH}")
